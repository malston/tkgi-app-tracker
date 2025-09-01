#!/usr/bin/env python3

"""
Excel Template Generator for TKGI Application Tracker
Generates Excel workbooks with pivot tables and charts for management analysis
"""

import os
import sys
import json
import csv
from pathlib import Path
from datetime import datetime
import argparse

try:
    import openpyxl
    from openpyxl.chart import PieChart, BarChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ExcelTemplateGenerator:
    """Generate Excel workbooks with pivot tables and charts"""

    def __init__(self, output_dir="reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

        self._ensure_excel_available()

    def _ensure_excel_available(self):
        """Ensure openpyxl is available for Excel generation"""
        global EXCEL_AVAILABLE
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl required for Excel generation. Install with: pip3 install -r requirements.txt")
            print("Or run 'make setup' to create a virtual environment with all dependencies.")
            sys.exit(1)

    def load_csv_data(self):
        """Load the latest CSV reports"""
        csv_files = {}

        # Find latest CSV files
        for report_type in ['application_report', 'executive_summary', 'migration_priority', 'cluster_report']:
            pattern = f"{report_type}_*.csv"
            files = list(self.output_dir.glob(pattern))
            if files:
                latest_file = max(files, key=lambda f: f.stat().st_mtime)
                csv_files[report_type] = latest_file
                print(f"Found {report_type}: {latest_file.name}")

        return csv_files

    def create_styled_workbook(self):
        """Create a new workbook with standard styling"""
        wb = openpyxl.Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        return wb

    def apply_header_style(self, ws, row_num=1):
        """Apply consistent header styling to a worksheet"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def apply_data_style(self, ws, start_row=2):
        """Apply consistent data styling"""
        thin_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")

    def create_application_sheet(self, wb, csv_file):
        """Create application data sheet with formatting"""
        ws = wb.create_sheet("Applications")

        # Load CSV data
        with open(csv_file, 'r') as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    # Convert numeric values
                    try:
                        if value.isdigit():
                            value = int(value)
                        elif value.replace('.', '').isdigit():
                            value = float(value)
                    except:
                        pass

                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Apply styling
        self.apply_header_style(ws)
        self.apply_data_style(ws)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        return ws

    def create_executive_dashboard(self, wb, csv_file):
        """Create executive dashboard with summary metrics"""
        ws = wb.create_sheet("Executive Dashboard")

        # Load executive summary data - proper CSV parsing for actual format
        metrics = {}
        foundation_data = {}

        with open(csv_file, 'r') as f:
            reader = csv.reader(f)
            current_section = "main"
            metrics[current_section] = {}

            for row in reader:
                if not row or (len(row) >= 1 and row[0] == ""):
                    continue

                if len(row) >= 2:
                    # Skip header row
                    if row[0] == "Metric" and row[1] == "Value":
                        continue
                    elif row[0] == "Foundation Breakdown":
                        current_section = "foundations"
                        continue
                    elif row[0] == "Migration Readiness":
                        current_section = "migration"
                        metrics[current_section] = {}
                        continue
                    elif row[0] == "Foundation" and len(row) >= 4:  # Foundation header row
                        continue
                    elif current_section == "foundations" and len(row) >= 4:
                        # Foundation data rows
                        foundation_name = row[0]
                        foundation_data[foundation_name] = {
                            'total': int(row[1]) if row[1].isdigit() else 0,
                            'active': int(row[2]) if row[2].isdigit() else 0,
                            'inactive': int(row[3]) if row[3].isdigit() else 0
                        }
                    else:
                        # Regular metric rows
                        try:
                            value = int(row[1]) if row[1].isdigit() else row[1]
                        except:
                            value = row[1]
                        metrics[current_section][row[0]] = value

        # Create dashboard layout
        row = 1

        # Title
        ws.cell(row=row, column=1, value="TKGI Application Tracker - Executive Dashboard")
        ws.cell(row=row, column=1).font = Font(size=16, bold=True, color="366092")
        ws.merge_cells(f'A{row}:H{row}')
        row += 2

        # Key metrics section
        ws.cell(row=row, column=1, value="Key Metrics")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        row += 1

        # Add key metrics in a grid
        col = 1
        for section, data in metrics.items():
            if section in ['main', 'migration']:
                for metric, value in data.items():
                    if metric not in ['Report Date']:
                        ws.cell(row=row, column=col, value=metric)
                        ws.cell(row=row+1, column=col, value=value)

                        # Style metric boxes
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                        ws.cell(row=row, column=col).font = Font(bold=True)
                        ws.cell(row=row+1, column=col).font = Font(size=12, bold=True, color="366092")

                        col += 1
                        if col > 6:  # Wrap to next row
                            col = 1
                            row += 3

        row += 3

        # Foundation breakdown table
        ws.cell(row=row, column=1, value="Foundation Breakdown")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        row += 1

        # Foundation table headers
        headers = ['Foundation', 'Total Apps', 'Active', 'Inactive', 'Active %']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)

        self.apply_header_style(ws, row)
        row += 1

        # Foundation data
        if foundation_data:
            for foundation_name, data in foundation_data.items():
                ws.cell(row=row, column=1, value=foundation_name)
                ws.cell(row=row, column=2, value=data['total'])
                ws.cell(row=row, column=3, value=data['active'])
                ws.cell(row=row, column=4, value=data['inactive'])
                if data['total'] > 0:
                    ws.cell(row=row, column=5, value=f"=C{row}/B{row}")  # Formula for percentage
                    ws.cell(row=row, column=5).number_format = '0%'
                else:
                    ws.cell(row=row, column=5, value=0)
                row += 1

        # Apply data styling to foundation table
        self.apply_data_style(ws, row - 3)

        # Auto-adjust columns
        for col in range(1, 6):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        return ws

    def create_charts_sheet(self, wb, apps_sheet):
        """Create charts and visualizations sheet with real data"""
        ws = wb.create_sheet("Charts & Analysis")

        # Title
        ws.cell(row=1, column=1, value="TKGI Application Analysis - Charts & Trends")
        ws.cell(row=1, column=1).font = Font(size=16, bold=True, color="366092")
        ws.merge_cells('A1:H1')

        # First, let's determine the data structure from the apps sheet
        app_data_rows = apps_sheet.max_row

        # Find column indices for key fields
        header_row = []
        for cell in apps_sheet[1]:
            header_row.append(cell.value)

        try:
            status_col = header_row.index('Status') + 1 if 'Status' in header_row else None
            environment_col = header_row.index('Environment') + 1 if 'Environment' in header_row else None
            readiness_col = header_row.index('Migration Readiness Score') + 1 if 'Migration Readiness Score' in header_row else None
        except ValueError:
            # Fallback column indices if headers don't match exactly
            status_col = 2  # Assume status is column B
            environment_col = 3  # Assume environment is column C
            readiness_col = 9  # Assume migration readiness is around column I

        # Create summary data for charts
        ws.cell(row=3, column=1, value="Chart Data (Auto-Generated)")
        ws.cell(row=3, column=1).font = Font(bold=True)

        # Status summary data
        ws.cell(row=5, column=1, value="Status")
        ws.cell(row=5, column=2, value="Count")
        ws.cell(row=6, column=1, value="Active")
        ws.cell(row=6, column=2, value="=COUNTIF(Applications!B:B,\"Active\")")
        ws.cell(row=7, column=1, value="Inactive")
        ws.cell(row=7, column=2, value="=COUNTIF(Applications!B:B,\"Inactive\")")

        # Environment summary data
        ws.cell(row=5, column=4, value="Environment")
        ws.cell(row=5, column=5, value="Count")
        ws.cell(row=6, column=4, value="Production")
        ws.cell(row=6, column=5, value="=COUNTIF(Applications!C:C,\"production\")")
        ws.cell(row=7, column=4, value="Non-Production")
        ws.cell(row=7, column=5, value="=COUNTIF(Applications!C:C,\"nonprod\")")
        ws.cell(row=8, column=4, value="Lab")
        ws.cell(row=8, column=5, value="=COUNTIF(Applications!C:C,\"lab\")")

        # Migration readiness ranges
        ws.cell(row=5, column=7, value="Readiness Range")
        ws.cell(row=5, column=8, value="Count")
        ws.cell(row=6, column=7, value="Ready (80-100)")
        ws.cell(row=6, column=8, value="=COUNTIFS(Applications!I:I,\">=80\",Applications!I:I,\"<=100\")")
        ws.cell(row=7, column=7, value="Planning (60-79)")
        ws.cell(row=7, column=8, value="=COUNTIFS(Applications!I:I,\">=60\",Applications!I:I,\"<80\")")
        ws.cell(row=8, column=7, value="Complex (40-59)")
        ws.cell(row=8, column=8, value="=COUNTIFS(Applications!I:I,\">=40\",Applications!I:I,\"<60\")")
        ws.cell(row=9, column=7, value="High Risk (0-39)")
        ws.cell(row=9, column=8, value="=COUNTIFS(Applications!I:I,\">=0\",Applications!I:I,\"<40\")")

        # Application Status Pie Chart
        chart1 = PieChart()
        chart1.title = "Application Status Distribution"

        data = Reference(ws, min_col=2, min_row=5, max_row=7, max_col=2)
        labels = Reference(ws, min_col=1, min_row=6, max_row=7)

        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(labels)
        chart1.width = 10
        chart1.height = 8

        ws.add_chart(chart1, "A11")

        # Environment Distribution Pie Chart
        chart2 = PieChart()
        chart2.title = "Applications by Environment"

        data = Reference(ws, min_col=5, min_row=5, max_row=8, max_col=5)
        labels = Reference(ws, min_col=4, min_row=6, max_row=8)

        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(labels)
        chart2.width = 10
        chart2.height = 8

        ws.add_chart(chart2, "H11")

        # Migration Readiness Bar Chart
        chart3 = BarChart()
        chart3.title = "Migration Readiness Distribution"
        chart3.x_axis.title = "Readiness Categories"
        chart3.y_axis.title = "Number of Applications"

        data = Reference(ws, min_col=8, min_row=5, max_row=9, max_col=8)
        labels = Reference(ws, min_col=7, min_row=6, max_row=9)

        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(labels)
        chart3.width = 12
        chart3.height = 8

        ws.add_chart(chart3, "A27")

        return ws

    def create_pivot_instructions(self, wb):
        """Create instructions sheet for pivot table creation"""
        ws = wb.create_sheet("Pivot Table Instructions")

        instructions = [
            "TKGI Application Tracker - Pivot Table Guide",
            "",
            "This workbook contains TKGI application data optimized for Excel analysis.",
            "Follow these instructions to create powerful pivot tables and reports:",
            "",
            "QUICK START PIVOT TABLES:",
            "",
            "1. APPLICATION SUMMARY BY ENVIRONMENT:",
            "   - Select data from 'Applications' sheet",
            "   - Insert > PivotTable",
            "   - Drag 'Environment' to Rows",
            "   - Drag 'Status' to Columns",
            "   - Drag 'Application ID' to Values (Count)",
            "",
            "2. MIGRATION READINESS ANALYSIS:",
            "   - Use 'Applications' sheet data",
            "   - Drag 'Migration Readiness Score' to Rows (Group by ranges: 0-39, 40-59, 60-79, 80-100)",
            "   - Drag 'Environment' to Columns",
            "   - Drag 'Application ID' to Values",
            "",
            "3. FOUNDATION UTILIZATION:",
            "   - Drag 'Foundations' to Rows",
            "   - Drag 'Total Pods' to Values (Sum)",
            "   - Drag 'Running Pods' to Values (Sum)",
            "   - Add calculated field: Pod Utilization = Running Pods / Total Pods",
            "",
            "4. INACTIVE APPLICATION ANALYSIS:",
            "   - Filter 'Status' = 'Inactive'",
            "   - Drag 'Days Since Activity' to Rows (Group: 0-30, 31-60, 61-90, 90+)",
            "   - Drag 'Environment' to Columns",
            "   - Identify candidates for decommissioning",
            "",
            "RECOMMENDED CHARTS:",
            "",
            "• Pie Chart: Application Status (Active vs Inactive)",
            "• Bar Chart: Applications by Foundation",
            "• Line Chart: Migration Readiness Trends (if historical data available)",
            "• Scatter Plot: Pod Count vs Migration Readiness Score",
            "",
            "DATA REFRESH:",
            "",
            "When new weekly reports are generated:",
            "1. Copy new CSV data to respective sheets",
            "2. Right-click any pivot table > Refresh",
            "3. Charts will automatically update",
            "",
            "FILTERS & SLICERS:",
            "",
            "Add slicers for interactive filtering:",
            "• Environment (Production/Non-Production)",
            "• Foundation (DC01/DC02/DC03/DC04)",
            "• Status (Active/Inactive)",
            "• Migration Readiness Score ranges",
            "",
            "CONDITIONAL FORMATTING:",
            "",
            "Apply to highlight key insights:",
            "• Red: Inactive apps > 90 days",
            "• Yellow: Migration score 40-69",
            "• Green: Migration score 70+",
            "• Blue: Production applications"
        ]

        for row, instruction in enumerate(instructions, 1):
            ws.cell(row=row, column=1, value=instruction)
            if row == 1:
                ws.cell(row=row, column=1).font = Font(size=16, bold=True, color="366092")
            elif instruction.startswith(("QUICK START", "RECOMMENDED", "DATA REFRESH", "FILTERS", "CONDITIONAL")):
                ws.cell(row=row, column=1).font = Font(size=12, bold=True, color="366092")
            elif instruction and not instruction.startswith(" "):
                ws.cell(row=row, column=1).font = Font(bold=True)

        # Auto-adjust column width
        ws.column_dimensions['A'].width = 100

        return ws

    def create_trend_analysis_template(self, wb):
        """Create trend analysis with historical data from JSON files"""
        ws = wb.create_sheet("Trend Analysis")

        # Headers for trend tracking
        headers = [
            "Week Ending", "Total Applications", "Active Applications", "Inactive Applications",
            "Ready for Migration", "Migrations Completed", "Migration Rate %"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self.apply_header_style(ws, 1)

        # Try to load historical data from JSON
        historical_data = self.load_historical_data()

        if historical_data:
            # Use actual historical data
            for idx, week_data in enumerate(historical_data, 2):
                ws.cell(row=idx, column=1, value=week_data.get('week_ending', ''))
                ws.cell(row=idx, column=2, value=week_data.get('total_applications', 0))
                ws.cell(row=idx, column=3, value=week_data.get('active_applications', 0))
                ws.cell(row=idx, column=4, value=week_data.get('total_applications', 0) - week_data.get('active_applications', 0))
                ws.cell(row=idx, column=5, value=week_data.get('ready_for_migration', 0))
                ws.cell(row=idx, column=6, value=week_data.get('migrations_completed_total', 0))

                # Migration rate calculation
                if idx > 2:  # Have previous week data
                    ws.cell(row=idx, column=7, value=f"=IF(C{idx-1}=0,0,F{idx}/C{idx-1})")
                    ws.cell(row=idx, column=7).number_format = '0.0%'
        else:
            # Fallback to template with formulas
            from datetime import datetime, timedelta
            for week in range(12):  # 12 weeks of template data
                row = week + 2
                week_date = datetime.now() - timedelta(weeks=11-week)

                ws.cell(row=row, column=1, value=week_date.strftime("%Y-%m-%d"))
                ws.cell(row=row, column=2, value=150 - week * 2)  # Sample declining trend
                ws.cell(row=row, column=3, value=int((150 - week * 2) * 0.68))  # ~68% active
                ws.cell(row=row, column=4, value=f"=B{row}-C{row}")  # Inactive
                ws.cell(row=row, column=5, value=int((150 - week * 2) * 0.41))  # ~41% ready
                ws.cell(row=row, column=6, value=week * 2)  # Cumulative migrations

                if week > 0:
                    ws.cell(row=row, column=7, value=f"=IF(C{row-1}=0,0,2/C{row-1})")  # 2 per week migration rate
                    ws.cell(row=row, column=7).number_format = '0.0%'

        # Add trend chart
        chart = LineChart()
        chart.title = "Application Migration Trends"
        chart.x_axis.title = "Week"
        chart.y_axis.title = "Application Count"

        # Reference data range
        data_rows = len(historical_data) + 1 if historical_data else 13
        data = Reference(ws, min_col=2, min_row=1, max_row=data_rows, max_col=6)
        categories = Reference(ws, min_col=1, min_row=2, max_row=data_rows)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.width = 15
        chart.height = 10

        ws.add_chart(chart, "I2")

        # Apply styling
        self.apply_data_style(ws, 2)

        # Auto-adjust columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        return ws

    def load_historical_data(self):
        """Load historical data from JSON files if available"""
        try:
            historical_files = list(self.output_dir.glob('historical_*.json'))
            if historical_files:
                latest_file = max(historical_files, key=lambda f: f.stat().st_mtime)
                with open(latest_file, 'r') as f:
                    import json
                    return json.load(f)
        except Exception as e:
            print(f"Could not load historical data: {e}")
        return None

    def generate_excel_workbook(self):
        """Generate complete Excel workbook with all sheets and features"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl required for Excel generation")

        print("Generating Excel workbook...")

        # Load CSV data
        csv_files = self.load_csv_data()

        if not csv_files:
            print("No CSV files found. Generate reports first.")
            return None

        # Create workbook
        wb = self.create_styled_workbook()

        # Create data sheets
        if 'application_report' in csv_files:
            apps_sheet = self.create_application_sheet(wb, csv_files['application_report'])
            print("✓ Created Applications data sheet")

        if 'executive_summary' in csv_files:
            self.create_executive_dashboard(wb, csv_files['executive_summary'])
            print("✓ Created Executive Dashboard")

        # Create analysis sheets
        if 'application_report' in csv_files:
            self.create_charts_sheet(wb, apps_sheet)
            print("✓ Created Charts & Analysis sheet")

        # Create instruction and template sheets
        self.create_pivot_instructions(wb)
        print("✓ Created Pivot Table Instructions")

        self.create_trend_analysis_template(wb)
        print("✓ Created Trend Analysis template")

        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = self.output_dir / f"TKGI_App_Tracker_Analysis_{timestamp}.xlsx"

        wb.save(excel_file)

        print(f"\n✓ Excel workbook saved: {excel_file}")
        print(f"  File size: {excel_file.stat().st_size / 1024 / 1024:.1f} MB")

        return excel_file

    def create_template_only(self):
        """Create template workbook without data for customization"""
        wb = self.create_styled_workbook()

        # Create empty template sheets
        self.create_pivot_instructions(wb)
        self.create_trend_analysis_template(wb)

        # Save template
        template_file = self.output_dir / "TKGI_App_Tracker_Template.xlsx"
        wb.save(template_file)

        print(f"✓ Excel template saved: {template_file}")
        return template_file

def main():
    parser = argparse.ArgumentParser(description='Generate Excel workbook for TKGI application analysis')
    parser.add_argument('-o', '--output-dir', default='reports',
                       help='Directory containing CSV files and for Excel output')
    parser.add_argument('--template-only', action='store_true',
                       help='Create template workbook without data')
    parser.add_argument('-v', '--verbose', action='store_true',
                       help='Verbose output')

    args = parser.parse_args()

    try:
        generator = ExcelTemplateGenerator(args.output_dir)

        if args.template_only:
            result = generator.create_template_only()
        else:
            result = generator.generate_excel_workbook()

        if result:
            print("\n" + "="*60)
            print("EXCEL WORKBOOK GENERATION COMPLETE")
            print("="*60)
            print(f"Generated: {result.name}")
            print("\nFeatures included:")
            print("  ✓ Application data with formatting")
            print("  ✓ Executive dashboard with key metrics")
            print("  ✓ Charts and visualizations")
            print("  ✓ Pivot table instructions")
            print("  ✓ Trend analysis template")
            print("  ✓ Professional styling and formatting")
            print("\nOpen in Excel and follow the 'Pivot Table Instructions' sheet")
            print("to create powerful analysis views.")
            print("="*60)

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
